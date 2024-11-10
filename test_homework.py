import pathlib
from io import TextIOWrapper
from zipfile import ZipFile, ZIP_DEFLATED
from pathlib import Path
import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import tmp_dir
import csv

folder = pathlib.Path(tmp_dir)
zip_file = Path('archive.zip').resolve()


@pytest.fixture()
def create_zip():
    with (ZipFile(zip_file, "w", ZIP_DEFLATED) as zip_archive):
        for file in folder.iterdir():
            zip_archive.write(file, file.name)
        for item in zip_archive.infolist():
            print(f"File Name: {item.filename} Date: {item.date_time} Size: {item.file_size}")


def test_pdf(create_zip):
    with (ZipFile(zip_file, "r", ZIP_DEFLATED) as zip_archive):
        pdf = zip_archive.open("Python Testing with Pytest (Brian Okken).pdf")
        zip_pdf = PdfReader(pdf)
        real_pdf = PdfReader("files/Python Testing with Pytest (Brian Okken).pdf")

        assert zip_pdf.pages[1].extract_text() == real_pdf.pages[1].extract_text()


def test_xlsx(create_zip):
    with (ZipFile(zip_file, "r", ZIP_DEFLATED) as zip_archive):
        xlsx = zip_archive.open("file_example_XLSX_50.xlsx")
        zip_xlsx = load_workbook(xlsx)

        real_xlsx = load_workbook("files/file_example_XLSX_50.xlsx")

        assert zip_xlsx.active.cell(row=2, column=3).value == real_xlsx.active.cell(row=2, column=3).value


def test_csv(create_zip):
    with (ZipFile(zip_file, "r", ZIP_DEFLATED) as zip_archive):
        csv_f = zip_archive.open("username.csv", 'r')
        zip_csv = csv.reader(TextIOWrapper(csv_f, 'utf-8'))

        real_csv = csv.reader(open("files/username.csv", 'r'))
        real_csv_list = list(real_csv)
        zip_csv_list = list(zip_csv)

        assert zip_csv_list[1:5] == real_csv_list[1:5]
